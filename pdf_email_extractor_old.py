# Enhanced PDF i Email Data Extractor wykorzystujący modele Hugging Face
# Autor: Rozszerzenie istniejącego rozwiązania o modele transformerowe

import pandas as pd
import PyPDF2
import pdfplumber
import re
import email
import imaplib
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging
from datetime import datetime
import nltk
import torch
from typing import List, Dict, Any, Tuple, Optional
import json

# Import bibliotek Hugging Face
from transformers import (
    AutoTokenizer, 
    AutoModelForTokenClassification,
    AutoModelForSequenceClassification,
    pipeline
)

# Konfiguracja logowania
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EnhancedDataExtractor:
    def __init__(self, use_transformers=True):
        self.use_transformers = use_transformers
        
        # Tradycyjne wzorce regex (jako fallback)
        self.patterns = {
            'customer_name': r'(?:Klant|Customer|Naam|voor):\s*([^\n\r]+)|(?:Hierbij de bestelling voor)\s*([^\n\r\.]+)',
            'po_number': r'(?:Boeknummer|PO|Order|BS|Orderreferentie)[\s:]*([A-Z0-9\-]+)',
            'material_code': r'(PPG\d+)',
            'material_description': r'(?:Sigma.*?(?:\d+\.?\d*\s*Ltr)|Opis.*?(?:\d+\.?\d*\s*Ltr)|Kit\s+\w+|Lakverf\s+\d+|Grondverf\s+\d+)',
            'shipping_street': r'(?:Adres|Address|Straat|Afleveradres):\s*([^\n\r,]+)',
            'shipping_postcode': r'(\d{4}\s*[A-Z]{2})',
            'colour_code': r'(?:RAL\s*\d+|No\.\s*\d+(?:\.\d+)?|\b\d{4}\b)',
            'fan_code': r'(?:Fan|Waaier):\s*([^\n\r]+)',
            'shipping_condition': r'(?:Verzending|Levering|Shipping|Ophaaldatum):\s*([^\n\r]+)',
            'project_number': r'(?:Project|Projectnummer|Projectdossier):\s*([^\n\r]+)',
            'date': r'(?:Datum|Ophaaldatum):\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
            'quantity': r'(\d+)\s*(?:stuks?|stuk|pieces?|pcs)\b',
            'reference_number': r'(?:Ref|Referentie|Referentienummer|Orderreferentie):\s*([^\n\r]+)'
        }
        
        if self.use_transformers:
            self._initialize_transformers()
        else:
            # Fallback do prostszego modelu klasyfikacji
            from sklearn.feature_extraction.text import TfidfVectorizer
            from sklearn.naive_bayes import MultinomialNB
            import pickle
            
            try:
                with open('email_classifier.pkl', 'rb') as f:
                    self.ml_model = pickle.load(f)
                with open('vectorizer.pkl', 'rb') as f:
                    self.vectorizer = pickle.load(f)
                logger.info("Załadowano istniejący model ML")
            except FileNotFoundError:
                logger.info("Trenowanie nowego modelu ML")
                self._train_simple_model()
    
    def _initialize_transformers(self):
        """Inicjalizuje modele transformerowe"""
        try:
            # Model do klasyfikacji dokumentów
            logger.info("Inicjalizacja modeli transformerowych...")
            
            # Można użyć mniejszych i szybszych modeli dla produkcji
            self.tokenizer_classifier = AutoTokenizer.from_pretrained("distilbert-base-uncased")
            self.classifier_model = AutoModelForSequenceClassification.from_pretrained("distilbert-base-uncased")
            
            # Model do rozpoznawania encji (NER)
            self.tokenizer_ner = AutoTokenizer.from_pretrained("dslim/bert-base-NER")
            self.ner_model = AutoModelForTokenClassification.from_pretrained("dslim/bert-base-NER")
            
            # Tworzenie pipeline'ów
            self.ner_pipeline = pipeline("ner", model=self.ner_model, tokenizer=self.tokenizer_ner)
            self.classifier_pipeline = pipeline("text-classification", model=self.classifier_model, tokenizer=self.tokenizer_classifier)
            
            logger.info("Modele transformerowe załadowane pomyślnie")
            
            # Opcjonalnie - fine-tuning na własnych danych
            self._fine_tune_if_needed()
            
        except Exception as e:
            logger.error(f"Błąd inicjalizacji modeli: {str(e)}")
            logger.warning("Przełączanie na tradycyjne metody ekstrakcji...")
            self.use_transformers = False
            self._train_simple_model()
    
    def _fine_tune_if_needed(self):
        """Fine-tuning modeli jeśli istnieją dane treningowe"""
        model_path = 'fine_tuned_ner_model'
        
        if os.path.exists(model_path):
            logger.info("Ładowanie dostrojonego modelu NER...")
            self.ner_model = AutoModelForTokenClassification.from_pretrained(model_path)
            self.ner_pipeline = pipeline("ner", model=self.ner_model, tokenizer=self.tokenizer_ner)
        else:
            logger.info("Brak dostrojonego modelu. Używanie modelu bazowego.")
            
            # Tutaj można dodać logikę do fine-tuningu na własnych danych
            # jeśli są dostępne odpowiednie dane treningowe
    
    def _train_simple_model(self):
        """Trenuje prosty model klasyfikacji jako fallback"""
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.naive_bayes import MultinomialNB
        import pickle
        
        # Przykładowe dane treningowe
        training_data = [
            ("bestelling order PPG materiaal", "order"),
            ("aanvraag offerte prijsopgave", "quote"),
            ("levering verzending adres", "delivery"),
            ("factuur betaling rekening", "invoice"),
            ("klacht probleem kwaliteit", "complaint"),
            ("Sigma Rapid Gloss bestelling", "order"),
            ("Kom ik ophalen", "pickup"),
            ("Afleveradres wijziging", "delivery"),
            ("BESTELLING", "order"),
            ("Boeknummer BS", "order"),
            ("Projectdossier", "order"),
            ("PPG Coatings", "order")
        ]
        
        texts, labels = zip(*training_data)
        
        self.vectorizer = TfidfVectorizer(stop_words='english', lowercase=True)
        X = self.vectorizer.fit_transform(texts)
        
        self.ml_model = MultinomialNB()
        self.ml_model.fit(X, labels)
        
        # Zapisz model
        with open('email_classifier.pkl', 'wb') as f:
            pickle.dump(self.ml_model, f)
        with open('vectorizer.pkl', 'wb') as f:
            pickle.dump(self.vectorizer, f)
    
    def extract_structured_info(self, text: str) -> Dict[str, Any]:
        """Ekstraktuje strukturalne informacje z wykorzystaniem NER"""
        if not self.use_transformers:
            return {}
        
        try:
            # Podziel tekst na krótsze fragmenty, jeśli jest zbyt długi
            # Modele transformerowe mają ograniczenie długości sekwencji
            max_length = 512  # Typowa maksymalna długość dla większości modeli BERT
            chunks = self._split_text_into_chunks(text, max_length)
            
            all_entities = []
            for chunk in chunks:
                chunk_entities = self.ner_pipeline(chunk)
                all_entities.extend(chunk_entities)
            
            # Filtrowanie i grupowanie encji
            structured_data = self._process_ner_results(all_entities, text)
            
            # Dodaj zaawansowaną analizę relacji i kontekstu
            structured_data = self._enhance_with_context(structured_data, text)
            
            return structured_data
        except Exception as e:
            logger.error(f"Błąd ekstrakcji strukturalnej: {str(e)}")
            return {}
    
    def _split_text_into_chunks(self, text: str, max_length: int) -> List[str]:
        """Dzieli tekst na mniejsze fragmenty dla modeli transformerowych"""
        words = text.split()
        chunks = []
        current_chunk = []
        current_length = 0
        
        for word in words:
            # Oszacujmy długość tokenów (w przybliżeniu)
            word_token_length = len(word.split()) + 1  # +1 dla spacji
            
            if current_length + word_token_length > max_length:
                chunks.append(" ".join(current_chunk))
                current_chunk = [word]
                current_length = word_token_length
            else:
                current_chunk.append(word)
                current_length += word_token_length
        
        if current_chunk:
            chunks.append(" ".join(current_chunk))
        
        return chunks
    
    def _process_ner_results(self, entities: List[Dict[str, Any]], original_text: str) -> Dict[str, Any]:
        """Przetwarza wyniki NER do struktury danych"""
        # Grupowanie encji według typów
        grouped_entities = {}
        for entity in entities:
            entity_type = entity.get('entity')
            if entity_type.startswith('B-'):  # Początek encji
                entity_type = entity_type[2:]  # Usuń prefiks B-
                
            if entity_type not in grouped_entities:
                grouped_entities[entity_type] = []
                
            grouped_entities[entity_type].append({
                'word': entity.get('word'),
                'score': entity.get('score'),
                'start': entity.get('start'),
                'end': entity.get('end')
            })
        
        # Mapowanie typów encji na pola dokumentu
        entity_mapping = {
            'ORG': 'customer_name',
            'PER': 'contact_person',
            'LOC': 'shipping_location',
            'MISC': 'miscellaneous',
            # Dodaj więcej mapowań w miarę potrzeb
        }
        
        structured_data = {}
        for entity_type, entities_list in grouped_entities.items():
            if entity_type in entity_mapping:
                field_name = entity_mapping[entity_type]
                # Wybierz encję z najwyższym wynikiem
                best_entity = max(entities_list, key=lambda x: x['score'])
                structured_data[field_name] = best_entity['word']
        
        # Dodatkowe przetwarzanie dla specyficznych pól
        structured_data = self._extract_custom_entities(structured_data, original_text)
        
        return structured_data
    
    def _extract_custom_entities(self, data: Dict[str, Any], text: str) -> Dict[str, Any]:
        """Ekstraktuje niestandardowe encje za pomocą reguł specyficznych dla dziedziny"""
        # Szukaj numerów zamówień (PO)
        po_matches = re.findall(r'(?:BS|Boeknummer)[\s:]*(\d+)', text, re.IGNORECASE)
        if po_matches:
            data['po_number'] = po_matches[0]
        
        # Szukaj kodów materiałów PPG (zachowaj wszystkie wystąpienia)
        ppg_matches = re.findall(r'PPG\d+', text)
        if ppg_matches:
            data['material_codes'] = ppg_matches  # Zachowaj wszystkie, łącznie z duplikatami
            if not data.get('material_code') and ppg_matches:
                data['material_code'] = ppg_matches[0]
        
        # Szukaj kodów pocztowych
        postcode_matches = re.findall(r'\b\d{4}\s*[A-Z]{2}\b', text)
        if postcode_matches:
            data['shipping_postcode'] = postcode_matches[0]
        
        # Szukaj dat
        date_matches = re.findall(r'\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b', text)
        if date_matches:
            data['date'] = date_matches[0]
        
        # Szukaj projektów
        project_matches = re.findall(r'(?:Project|Projectdossier)[\s:]*([^\n\r]+)', text, re.IGNORECASE)
        if project_matches:
            data['project_number'] = project_matches[0].strip()
        
        # Szukaj kodów kolorów RAL (zachowaj wszystkie wystąpienia)
        ral_matches = re.findall(r'(?:RAL\s*\d+|No\.\s*\d+(?:\.\d+)?)', text, re.IGNORECASE)
        if ral_matches:
            if not data.get('colour_codes'):
                data['colour_codes'] = ral_matches
            if not data.get('colour_code'):
                data['colour_code'] = ral_matches[0]
        
        # Szukaj ilości (bardziej precyzyjnie)
        # Szukaj wzorców w kontekście materiałów/produktów
        quantity_patterns = [
            r'(\d+)\s+(?:stuks?|stuk|pieces?|pcs)\b',  # Liczba + spacja + jednostka
            r'(?:stuks?|stuk|pieces?|pcs)[\s:]*(\d+)',  # Jednostka + liczba
            r'(?:aantal|quantity|qty)[\s:]*(\d+)',      # Słowo kluczowe + liczba
        ]
        
        all_quantities = []
        for pattern in quantity_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            all_quantities.extend(matches)
        
        # Filtruj nierealistyczne ilości (np. kody pocztowe, lata)
        filtered_quantities = []
        for qty in all_quantities:
            try:
                qty_int = int(qty)
                # Filtruj ilości które są prawdopodobnie błędne
                if 1 <= qty_int <= 10000:  # Realistyczny zakres ilości
                    filtered_quantities.append(qty)
            except ValueError:
                continue
        
        if filtered_quantities:
            if not data.get('quantities'):
                data['quantities'] = filtered_quantities
            if not data.get('quantity'):
                data['quantity'] = filtered_quantities[0]
        
        return data
    
    def _enhance_with_context(self, data: Dict[str, Any], text: str) -> Dict[str, Any]:
        """Dodaje kontekstowe informacje na podstawie analizy tekstu"""
        # Określenie typu dokumentu
        if "BESTELLING" in text.upper():
            data['document_type'] = 'order'
        elif "FACTUUR" in text.upper():
            data['document_type'] = 'invoice'
        elif "LEVERING" in text.upper():
            data['document_type'] = 'delivery'
        
        # Ekstrakcja pozycji zamówienia, jeśli to zamówienie
        if data.get('document_type') == 'order':
            order_items = self._extract_order_items(text)
            if order_items:
                data['order_items'] = order_items
        
        return data
    
    def _extract_order_items(self, text: str) -> List[Dict[str, Any]]:
        """Ekstraktuje pozycje zamówienia z tekstu z lepszą obsługą wielu pozycji"""
        items = []
        
        # Szukaj bloków tekstu zawierających informacje o produktach
        lines = text.split('\n')
        current_item = None
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
                
            # Szukaj linii zawierających kod PPG lub opis materiału
            if 'PPG' in line or any(keyword in line.upper() for keyword in ['SIGMA', 'RAPID', 'GLOSS']):
                # Zakończ poprzedni element jeśli istnieje
                if current_item and current_item.get('description'):
                    items.append(current_item)
                
                # Rozpocznij nowy element
                current_item = {
                    'code': '',
                    'description': line,
                    'quantity': '',
                    'color': '',
                    'unit': ''
                }
                
                # Ekstrakcja kodu PPG
                ppg_match = re.search(r'(PPG\d+)', line)
                if ppg_match:
                    current_item['code'] = ppg_match.group(1)
                
                # Szukaj ilości w tej samej linii - bardziej precyzyjne wzorce
                quantity_patterns = [
                    r'(\d+)\s+(?:stuks?|stuk|pieces?|pcs)\b',
                    r'(?:stuks?|stuk|pieces?|pcs)[\s:]*(\d+)',
                    r'(\d+)\s*(?:x\s*)?(?:ltr|litr|liter)',
                ]
                
                for pattern in quantity_patterns:
                    quantity_match = re.search(pattern, line, re.IGNORECASE)
                    if quantity_match:
                        qty = quantity_match.group(1)
                        # Sprawdź czy to realistyczna ilość
                        try:
                            qty_int = int(qty)
                            if 1 <= qty_int <= 10000:
                                current_item['quantity'] = qty
                                # Spróbuj znaleźć jednostkę
                                unit_match = re.search(r'(stuks?|stuk|pieces?|pcs|ltr|litr)', line, re.IGNORECASE)
                                if unit_match:
                                    current_item['unit'] = unit_match.group(1)
                                break
                        except ValueError:
                            continue
                
                # Ekstrakcja koloru
                color_match = re.search(r'(RAL\s*\d+|No\.\s*\d+)', line, re.IGNORECASE)
                if color_match:
                    current_item['color'] = color_match.group(1)
            
            elif current_item:
                # Sprawdź czy linia zawiera dodatkowe informacje o bieżącym elemencie
                if not current_item.get('quantity'):
                    for pattern in quantity_patterns:
                        quantity_match = re.search(pattern, line, re.IGNORECASE)
                        if quantity_match:
                            qty = quantity_match.group(1)
                            try:
                                qty_int = int(qty)
                                if 1 <= qty_int <= 10000:
                                    current_item['quantity'] = qty
                                    unit_match = re.search(r'(stuks?|stuk|pieces?|pcs|ltr|litr)', line, re.IGNORECASE)
                                    if unit_match:
                                        current_item['unit'] = unit_match.group(1)
                                    break
                            except ValueError:
                                continue
                
                if not current_item.get('color'):
                    color_match = re.search(r'(RAL\s*\d+|No\.\s*\d+)', line, re.IGNORECASE)
                    if color_match:
                        current_item['color'] = color_match.group(1)
        
        # Dodaj ostatni element jeśli istnieje
        if current_item and current_item.get('description'):
            items.append(current_item)
        
        return items
    
    def classify_document(self, text: str) -> Tuple[str, float]:
        """Klasyfikuje dokument za pomocą modelu transformerowego lub fallbacku"""
        if self.use_transformers:
            try:
                # Obetnij tekst, jeśli jest zbyt długi
                max_length = 512
                if len(text) > max_length:
                    text = text[:max_length]
                
                result = self.classifier_pipeline(text)[0]
                label = result['label']
                score = result['score']
                
                # Mapuj etykiety modelu na nasze kategorie
                label_mapping = {
                    'LABEL_0': 'order',
                    'LABEL_1': 'invoice',
                    'LABEL_2': 'delivery',
                    'LABEL_3': 'quote',
                    'LABEL_4': 'complaint'
                }
                
                return label_mapping.get(label, label), score
            except Exception as e:
                logger.error(f"Błąd klasyfikacji transformerowej: {str(e)}")
                # Fallback do tradycyjnej metody
                return self._classify_with_traditional_method(text)
        else:
            return self._classify_with_traditional_method(text)
    
    def _classify_with_traditional_method(self, text: str) -> Tuple[str, float]:
        """Klasyfikacja dokumentu za pomocą tradycyjnej metody"""
        try:
            # Upraszczamy tekst
            text_lower = text.lower()
            
            # Prosta klasyfikacja na podstawie słów kluczowych
            if 'bestelling' in text_lower or 'order' in text_lower:
                return 'order', 0.9
            elif 'factuur' in text_lower or 'invoice' in text_lower:
                return 'invoice', 0.9
            elif 'levering' in text_lower or 'delivery' in text_lower:
                return 'delivery', 0.9
            elif 'offerte' in text_lower or 'quote' in text_lower:
                return 'quote', 0.9
            
            # Użyj modelu ML jako ostatnia deska ratunku
            if hasattr(self, 'ml_model') and hasattr(self, 'vectorizer'):
                X = self.vectorizer.transform([text_lower])
                prediction = self.ml_model.predict(X)[0]
                confidence = max(self.ml_model.predict_proba(X)[0])
                return prediction, confidence
            
            return "unknown", 0.5
        except Exception as e:
            logger.error(f"Błąd klasyfikacji tradycyjnej: {str(e)}")
            return "unknown", 0.5
    
    def extract_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """Ekstraktuje dane z pliku PDF z wykorzystaniem transformerów"""
        extracted_data = {}
        
        try:
            # Odczytaj tekst z PDF
            text = self._extract_text_from_pdf(pdf_path)
            
            if not text.strip():
                logger.warning(f"Nie udało się wyekstrahować tekstu z PDF: {pdf_path}")
                return {}
            
            # Klasyfikuj dokument
            doc_type, confidence = self.classify_document(text)
            extracted_data['document_type'] = doc_type
            extracted_data['classification_confidence'] = confidence
            
            # Ekstrakcja danych za pomocą transformerów
            if self.use_transformers:
                transformer_data = self.extract_structured_info(text)
                extracted_data.update(transformer_data)
            
            # Dodaj ekstrakcję za pomocą regex jako uzupełnienie
            regex_data = self._extract_with_regex(text)
            
            # Łączenie danych - preferowane są dane z transformerów
            for key, value in regex_data.items():
                if key not in extracted_data or not extracted_data[key]:
                    extracted_data[key] = value
            
            # Dodatkowe przetwarzanie dla zamówień
            if doc_type == 'order' and not extracted_data.get('order_items'):
                order_items = self._extract_order_items(text)
                if order_items:
                    extracted_data['order_items'] = order_items
            
            logger.info(f"Pomyślnie wyekstraktowano dane z PDF: {pdf_path}")
            return extracted_data
            
        except Exception as e:
            logger.error(f"Błąd podczas ekstrakcji z PDF {pdf_path}: {str(e)}")
            return {}
    
    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        """Ekstraktuje tekst z pliku PDF"""
        text = ""
        
        # Próba z pdfplumber (lepsze dla tabel)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            logger.warning(f"Błąd ekstrakcji z pdfplumber: {str(e)}")
        
        # Fallback do PyPDF2 jeśli pdfplumber nie zadziałał
        if not text.strip():
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
            except Exception as e:
                logger.error(f"Błąd ekstrakcji z PyPDF2: {str(e)}")
        
        return text
    
    def _extract_with_regex(self, text: str) -> Dict[str, Any]:
        """Ekstraktuje dane za pomocą wyrażeń regularnych z obsługą wielu pozycji"""
        extracted_data = {}
        
        # Ekstrakcja danych z wykorzystaniem regex
        for field, pattern in self.patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
            if matches:
                if field == 'material_description':
                    # Specjalne przetwarzanie dla opisów materiałów
                    descriptions = []
                    for match in matches:
                        # Podziel każdy opis na części oddzielone średnikami
                        parts = [part.strip() for part in match.split(';') if part.strip()]
                        descriptions.extend(parts)
                    
                    if descriptions:
                        extracted_data[field] = "; ".join(descriptions)
                        # Dodatkowe pola dla lepszego przetwarzania
                        extracted_data['material_descriptions_list'] = descriptions
                elif field == 'material_code':
                    # Zbierz wszystkie kody materiałów (zachowaj duplikaty)
                    all_codes = matches  # Nie usuwaj duplikatów
                    extracted_data[field] = all_codes[0]
                    if len(all_codes) > 1:
                        extracted_data['material_codes'] = all_codes
                elif field == 'colour_code':
                    # Zbierz wszystkie kody kolorów (zachowaj duplikaty)
                    all_colours = matches  # Nie usuwaj duplikatów
                    extracted_data[field] = all_colours[0]
                    if len(all_colours) > 1:
                        extracted_data['colour_codes'] = all_colours
                elif field == 'quantity':
                    # Zbierz wszystkie ilości (zachowaj duplikaty)
                    all_quantities = matches  # Nie usuwaj duplikatów
                    extracted_data[field] = all_quantities[0]
                    if len(all_quantities) > 1:
                        extracted_data['quantities'] = all_quantities
                else:
                    extracted_data[field] = matches[0].strip()
        
        return extracted_data
    
    def extract_from_email(self, email_content: str) -> Dict[str, Any]:
        """Ekstraktuje dane z treści emaila z pełną obsługą zamówień"""
        extracted_data = {}
        
        # Klasyfikacja emaila
        email_type, confidence = self.classify_document(email_content)
        extracted_data['email_type'] = email_type
        extracted_data['confidence'] = confidence
        
        # Ekstrakcja za pomocą transformerów
        if self.use_transformers:
            transformer_data = self.extract_structured_info(email_content)
            extracted_data.update(transformer_data)
        
        # Uzupełnienie za pomocą regex
        regex_data = self._extract_with_regex(email_content)
        for key, value in regex_data.items():
            if key not in extracted_data or not extracted_data[key]:
                extracted_data[key] = value
        
        # Dodatkowe przetwarzanie specyficzne dla emaili
        extracted_data = self._enhance_email_extraction(extracted_data, email_content)
        
        # Specjalne przetwarzanie dla różnych typów emaili
        if email_type == 'order':
            # Dodatkowe przetwarzanie dla zamówień w emailach
            if 'ophalen' in email_content.lower():
                extracted_data['shipping_condition'] = 'Pickup/Afhalen'
            
            # Ekstrakcja pozycji zamówienia z emaila
            if not extracted_data.get('order_items'):
                order_items = self._extract_order_items(email_content)
                if order_items:
                    extracted_data['order_items'] = order_items
        
        logger.info(f"Wyekstraktowano dane z emaila (typ: {email_type}, pewność: {confidence:.2f})")
        return extracted_data

    def _enhance_email_extraction(self, data: Dict[str, Any], email_content: str) -> Dict[str, Any]:
        """Ulepszona ekstrakcja danych specyficzna dla emaili"""
        
        # Szukaj informacji o kliencie w emailach (często w podpisie lub nagłówku)
        if not data.get('customer_name'):
            # Różne wzorce dla nazw klientów w emailach
            customer_patterns = [
                r'(?:Van|From|Od):\s*([^\n\r<@]+?)(?:\s*<|@|\n)',
                r'(?:Met vriendelijke groet|Best regards|Groeten),?\s*\n\s*([^\n\r]+)',
                r'(?:Namens|Voor|On behalf of):\s*([^\n\r]+)',
            ]
            
            for pattern in customer_patterns:
                match = re.search(pattern, email_content, re.IGNORECASE)
                if match:
                    customer_name = match.group(1).strip()
                    # Filtruj oczywiste błędy (adresy email, cyfry)
                    if not re.search(r'[@\d]{3,}', customer_name) and len(customer_name) > 2:
                        data['customer_name'] = customer_name
                        break
        
        # Szukaj informacji o adresie dostawy w emailach
        if not data.get('shipping_street'):
            address_patterns = [
                r'(?:Afleveradres|Delivery address|Adres|Address)[:]\s*([^\n\r]+)',
                r'(?:Bezorgen naar|Deliver to)[:]\s*([^\n\r]+)',
                r'(?:Graag leveren op|Please deliver to)[:]\s*([^\n\r]+)',
            ]
            
            for pattern in address_patterns:
                match = re.search(pattern, email_content, re.IGNORECASE)
                if match:
                    data['shipping_street'] = match.group(1).strip()
                    break
        
        # Szukaj materiałów opisanych w różny sposób w emailach
        if not data.get('material_description') or not data.get('material_descriptions_list'):
            # Emaile mogą zawierać listy materiałów w różnych formatach
            material_patterns = [
                r'(?:Ik wil graag bestellen|I would like to order)[:]\s*\n?(.*?)(?:\n\n|\nMet vriendelijke|$)',
                r'(?:Bestelling|Order)[:]\s*\n?(.*?)(?:\n\n|\nGroeten|$)',
                r'(?:Materialen|Materials)[:]\s*\n?(.*?)(?:\n\n|\n[A-Z]|$)',
            ]
            
            for pattern in material_patterns:
                match = re.search(pattern, email_content, re.IGNORECASE | re.DOTALL)
                if match:
                    material_text = match.group(1).strip()
                    
                    # Podziel na linie i wyczyść
                    material_lines = []
                    for line in material_text.split('\n'):
                        line = line.strip()
                        if line and not line.startswith(('Met vriendelijke', 'Best regards', 'Groeten')):
                            # Usuń numerację jeśli istnieje (1., 2., -, *)
                            line = re.sub(r'^\d+\.?\s*[-*•]?\s*', '', line)
                            if line:  # Jeśli coś zostało po czyszczeniu
                                material_lines.append(line)
                    
                    if material_lines:
                        data['material_description'] = '; '.join(material_lines)
                        data['material_descriptions_list'] = material_lines
                        break
        
        # Szukaj kodów PPG w całym emailu (mogą być w różnych miejscach)
        ppg_codes = re.findall(r'PPG\d+', email_content, re.IGNORECASE)
        if ppg_codes:
            data['material_codes'] = ppg_codes
            if not data.get('material_code'):
                data['material_code'] = ppg_codes[0]
        
        # Szukaj ilości w kontekście emaila
        if not data.get('quantity') and not data.get('quantities'):
            # W emailach ilości mogą być wyrażone różnie
            quantity_patterns = [
                r'(\d+)\s*(?:stuks?|stuk|pieces?|pcs|x)\b',
                r'(?:aantal|quantity|qty)[:]\s*(\d+)',
                r'(\d+)\s*(?:liter|ltr|litr)',
                r'(\d+)\s*(?:buckets?|emmers?)',
            ]
            
            all_quantities = []
            for pattern in quantity_patterns:
                matches = re.findall(pattern, email_content, re.IGNORECASE)
                for match in matches:
                    try:
                        qty_int = int(match)
                        if 1 <= qty_int <= 10000:  # Realistyczny zakres
                            all_quantities.append(match)
                    except ValueError:
                        continue
            
            if all_quantities:
                data['quantities'] = all_quantities
                data['quantity'] = all_quantities[0]
        
        return data
    def read_emails_from_imap(self, server: str, username: str, password: str, folder: str = 'INBOX') -> List[Dict[str, Any]]:
        """Odczytuje emaile z serwera IMAP"""
        emails_data = []
        
        try:
            mail = imaplib.IMAP4_SSL(server)
            mail.login(username, password)
            mail.select(folder)
            
            # Szukaj emaili (możesz dostosować kryteria)
            _, message_numbers = mail.search(None, 'ALL')
            
            for num in message_numbers[0].split():
                _, msg_data = mail.fetch(num, '(RFC822)')
                email_body = msg_data[0][1]
                email_message = email.message_from_bytes(email_body)
                
                # Wyciągnij treść emaila
                content = ""
                if email_message.is_multipart():
                    for part in email_message.walk():
                        if part.get_content_type() == "text/plain":
                            try:
                                part_content = part.get_payload(decode=True)
                                if part_content:
                                    content += part_content.decode('utf-8', errors='ignore')
                            except Exception as e:
                                logger.warning(f"Błąd dekodowania części emaila: {str(e)}")
                else:
                    try:
                        payload = email_message.get_payload(decode=True)
                        if payload:
                            content += payload.decode('utf-8', errors='ignore')
                    except Exception as e:
                        logger.warning(f"Błąd dekodowania emaila: {str(e)}")
                
                # Ekstraktuj dane
                extracted = self.extract_from_email(content)
                extracted['email_subject'] = email_message['Subject']
                extracted['email_from'] = email_message['From']
                extracted['email_date'] = email_message['Date']
                
                emails_data.append(extracted)
            
            mail.close()
            mail.logout()
            
        except Exception as e:
            logger.error(f"Błąd podczas odczytu emaili: {str(e)}")
        
        return emails_data
    
    def save_to_excel(self, data_list: List[Dict[str, Any]], output_file: str) -> bool:
        """Zapisuje wyekstraktowane dane do pliku Excel z osobnymi wierszami dla każdego materiału"""
        try:
            # Przygotuj dane dla DataFrame
            rows = []
            for data in data_list:
                # Sprawdź czy material_description zawiera wiele pozycji
                material_descriptions = []
                if data.get('material_description'):
                    # Podziel opisy materiałów używając ";" jako separatora
                    material_descriptions = [desc.strip() for desc in data.get('material_description', '').split(';') if desc.strip()]
                
                # Jeśli brak opisów materiałów, stwórz jeden wiersz z pustym opisem
                if not material_descriptions:
                    material_descriptions = ['']
                
                # Obsługa order_items jako JSON string jeśli to lista
                order_items_str = ""
                if 'order_items' in data:
                    if isinstance(data['order_items'], list):
                        order_items_str = json.dumps(data['order_items'], ensure_ascii=False)
                    else:
                        order_items_str = str(data['order_items'])
                
                # Stwórz osobny wiersz dla każdego opisu materiału
                for i, material_desc in enumerate(material_descriptions):
                    # Sprawdź czy istnieją odpowiadające kody materiałów
                    material_code = ""
                    if data.get('material_codes') and isinstance(data['material_codes'], list):
                        # Jeśli mamy listę kodów, użyj odpowiadającego kodu (jeśli istnieje)
                        if i < len(data['material_codes']):
                            material_code = data['material_codes'][i]
                        elif len(data['material_codes']) == 1:
                            # Jeśli jest tylko jeden kod dla wielu opisów, użyj go dla wszystkich
                            material_code = data['material_codes'][0]
                    elif data.get('material_code'):
                        # Użyj pojedynczego kodu materiału
                        material_code = data['material_code']
                    
                    # Sprawdź czy istnieją odpowiadające kody kolorów
                    colour_code = ""
                    if data.get('colour_codes') and isinstance(data['colour_codes'], list):
                        if i < len(data['colour_codes']):
                            colour_code = data['colour_codes'][i]
                    elif data.get('colour_code'):
                        colour_code = data['colour_code']
                    
                    # Sprawdź czy istnieją odpowiadające ilości
                    quantity = ""
                    if data.get('quantities') and isinstance(data['quantities'], list):
                        if i < len(data['quantities']):
                            quantity = data['quantities'][i]
                    elif data.get('quantity'):
                        quantity = data['quantity']
                    
                    row = {
                        # Pola zielone (wymagane)
                        'Customer_Name': data.get('customer_name', ''),
                        'PO_Number': data.get('po_number', ''),
                        'Material_Code': material_code,
                        'Material_Description': material_desc,
                        'Shipping_Street': data.get('shipping_street', ''),
                        'Shipping_Postcode': data.get('shipping_postcode', ''),
                        'Colour_Code': colour_code,
                        'Fan_Code': data.get('fan_code', ''),
                        'Shipping_Condition': data.get('shipping_condition', ''),
                        
                        # Pola dodatkowe
                        'Quantity': quantity,
                        'Project_Number': data.get('project_number', ''),
                        'Date': data.get('date', ''),
                        'Document_Type': data.get('document_type', ''),
                        'Email_Type': data.get('email_type', ''),
                        'Confidence': data.get('confidence', ''),
                        'Reference_Number': data.get('reference_number', ''),
                        'Order_Items': order_items_str if i == 0 else '',  # Pokazuj order_items tylko w pierwszym wierszu
                        'Source_File': data.get('source_file', ''),
                        'Processing_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Extraction_Method': 'Transformer' if self.use_transformers else 'Traditional',
                        'Item_Number': i + 1,  # Numer pozycji w ramach jednego dokumentu
                        'Total_Items': len(material_descriptions)  # Łączna liczba pozycji w dokumencie
                    }
                    rows.append(row)
            
            # Utwórz DataFrame
            df = pd.DataFrame(rows)
            
            # Zapisz do Excel z formatowaniem
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Extracted_Data', index=False)
                
                # Dodaj formatowanie
                workbook = writer.book
                worksheet = writer.sheets['Extracted_Data']
                
                # Kolor dla nagłówków
                green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                
                # Formatowanie nagłówków
                for i, column in enumerate(df.columns):
                    cell = worksheet.cell(row=1, column=i+1)
                    
                    # Pola wymagane (zielone)
                    required_fields = ['Customer_Name', 'PO_Number', 'Material_Code', 
                                    'Material_Description', 'Shipping_Street', 
                                    'Shipping_Postcode', 'Colour_Code', 
                                    'Fan_Code', 'Shipping_Condition']
                    
                    # Nowe pola pomocnicze (żółte)
                    helper_fields = ['Item_Number', 'Total_Items', 'Quantity']
                    
                    if column in required_fields:
                        cell.fill = green_fill
                    elif column in helper_fields:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = light_blue_fill
                
                # Dostosuj szerokość kolumn
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Dane zapisane do pliku Excel: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Błąd podczas zapisywania do Excel: {str(e)}")
            return False
    
    def process_directory(self, directory_path: str, output_file: str) -> bool:
        """Przetwarza wszystkie pliki PDF w katalogu"""
        all_data = []
        
        for filename in os.listdir(directory_path):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(directory_path, filename)
                extracted_data = self.extract_from_pdf(pdf_path)
                extracted_data['source_file'] = filename
                all_data.append(extracted_data)
        
        if all_data:
            return self.save_to_excel(all_data, output_file)
        else:
            logger.warning("Nie znaleziono danych do przetworzenia")
            return False
    
    def evaluate_extraction_quality(self, extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """Ocena jakości ekstrakcji danych"""
        quality_metrics = {
            'completeness': 0.0,
            'confidence': 0.0,
            'required_fields_present': False,
            'warnings': []
        }
        
        # Lista wymaganych pól
        required_fields = ['customer_name', 'po_number', 'material_code', 
                          'shipping_street', 'shipping_postcode']
        
        # Sprawdź kompletność
        present_fields = sum(1 for field in required_fields if field in extracted_data and extracted_data[field])
        quality_metrics['completeness'] = present_fields / len(required_fields)
        
        # Sprawdź czy wszystkie wymagane pola są obecne
        quality_metrics['required_fields_present'] = all(field in extracted_data and extracted_data[field] 
                                                         for field in required_fields)
        
        # Średnia pewność dla pól ekstrakcji
        confidences = []
        if 'confidence' in extracted_data:
            confidences.append(float(extracted_data['confidence']))
        if 'classification_confidence' in extracted_data:
            confidences.append(float(extracted_data['classification_confidence']))
            
        if confidences:
            quality_metrics['confidence'] = sum(confidences) / len(confidences)
        
        # Ostrzeżenia
        for field in required_fields:
            if field not in extracted_data or not extracted_data[field]:
                quality_metrics['warnings'].append(f"Brak wymaganego pola: {field}")
        
        # Sprawdź wiarygodność danych (np. format kodu pocztowego)
        if 'shipping_postcode' in extracted_data:
            postcode = extracted_data['shipping_postcode']
            if not re.match(r'\d{4}\s*[A-Z]{2}', postcode):
                quality_metrics['warnings'].append(f"Niepoprawny format kodu pocztowego: {postcode}")
        
        return quality_metrics
    
    def export_to_json(self, data_list: List[Dict[str, Any]], output_file: str) -> bool:
        """Eksportuje dane do pliku JSON"""
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data_list, f, indent=4, ensure_ascii=False)
            logger.info(f"Dane zapisane do pliku JSON: {output_file}")
            return True
        except Exception as e:
            logger.error(f"Błąd podczas zapisywania do JSON: {str(e)}")
            return False

# Funkcja główna
def main():
    """Główna funkcja programu"""
    print("=== Enhanced PDF i Email Data Extractor ===")
    print("Inicjalizacja ekstraktora danych...")
    
    # Sprawdź czy dostępne są modele transformerowe
    use_transformers = True
    try:
        import torch
        from transformers import AutoTokenizer
        print("Dostępne biblioteki transformers. Używanie zaawansowanych modeli...")
    except ImportError:
        use_transformers = False
        print("Brak bibliotek transformers. Używanie tradycyjnych metod ekstrakcji...")
    
    extractor = EnhancedDataExtractor(use_transformers=use_transformers)
    
    # Konfiguracja - dostosuj do swoich potrzeb
    PDF_DIRECTORY = "pdfs"  # Katalog z plikami PDF
    OUTPUT_FILE = "extracted_data.xlsx"
    
    # Konfiguracja emaili (opcjonalnie)
    EMAIL_CONFIG = {
        'server': 'imap.gmail.com',  # Dostosuj do swojego dostawcy
        'username': 'your_email@gmail.com',
        'password': 'your_password',  # Użyj hasła aplikacji dla Gmail
        'folder': 'INBOX'
    }
    
    print("\nWybierz operację:")
    print("1. Przetwarzaj pliki PDF z katalogu")
    print("2. Przetwarzaj emaile z serwera IMAP")
    print("3. Przetwarzaj pojedynczy plik PDF")
    print("4. Przetwarzaj pojedynczy email (tekst)")
    print("5. Trenuj model ML")
    print("6. Eksportuj dane do JSON")
    print("7. Ewaluacja jakości ekstrakcji")
    
    choice = input("Wybierz opcję (1-7): ")
    
    if choice == '1':
        # Przetwarzanie katalogów PDF
        if not os.path.exists(PDF_DIRECTORY):
            os.makedirs(PDF_DIRECTORY)
            print(f"Utworzono katalog {PDF_DIRECTORY}. Umieść w nim pliki PDF i uruchom ponownie.")
            return
        
        success = extractor.process_directory(PDF_DIRECTORY, OUTPUT_FILE)
        if success:
            print(f"Przetwarzanie zakończone. Wyniki zapisane w {OUTPUT_FILE}")
        else:
            print("Błąd podczas przetwarzania")
    
    elif choice == '2':
        # Przetwarzanie emaili
        print("Podaj dane dostępowe do emaila:")
        EMAIL_CONFIG['server'] = input("Serwer IMAP (np. imap.gmail.com): ") or EMAIL_CONFIG['server']
        EMAIL_CONFIG['username'] = input("Email: ")
        EMAIL_CONFIG['password'] = input("Hasło: ")
        
        emails_data = extractor.read_emails_from_imap(**EMAIL_CONFIG)
        
        if emails_data:
            success = extractor.save_to_excel(emails_data, OUTPUT_FILE)
            if success:
                print(f"Przetwarzanie emaili zakończone. Wyniki zapisane w {OUTPUT_FILE}")
                print(f"Przetworzono {len(emails_data)} emaili")
                
                # Pokaż statystyki
                total_items = 0
                for email_data in emails_data:
                    if email_data.get('material_descriptions_list'):
                        total_items += len(email_data['material_descriptions_list'])
                    else:
                        total_items += 1
                print(f"Łączna liczba pozycji w Excel: {total_items}")
        else:
            print("Nie znaleziono emaili do przetworzenia")
    
    elif choice == '4':
        # Przetwarzanie pojedynczego emaila z pliku tekstowego
        txt_path = input("Podaj ścieżkę do pliku tekstowego z emailem: ").strip()
        
        if not txt_path:
            print("Nie podano ścieżki do pliku")
        elif not os.path.exists(txt_path):
            print(f"Plik nie istnieje: {txt_path}")
        else:
            try:
                # Wczytaj treść z pliku tekstowego
                with open(txt_path, 'r', encoding='utf-8') as f:
                    email_content = f.read()
                
                if email_content.strip():
                    print(f"Wczytano treść emaila z pliku: {txt_path}")
                    print(f"Długość tekstu: {len(email_content)} znaków")
                    
                    extracted_data = extractor.extract_from_email(email_content)
                    extracted_data['source_file'] = os.path.basename(txt_path)
                    
                    # Pokaż statystyki jakości
                    quality = extractor.evaluate_extraction_quality(extracted_data)
                    print(f"\nJakość ekstrakcji:")
                    print(f"- Kompletność: {quality['completeness']*100:.1f}%")
                    print(f"- Średnia pewność: {quality['confidence']*100:.1f}%")
                    print(f"- Wszystkie wymagane pola: {'Tak' if quality['required_fields_present'] else 'Nie'}")
                    
                    if quality['warnings']:
                        print("Ostrzeżenia:")
                        for warning in quality['warnings']:
                            print(f"  - {warning}")
                    
                    # Wygeneruj nazwę pliku wyjściowego z timestampem
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    output_file_name = f"email_from_file_{timestamp}.xlsx"
                    
                    success = extractor.save_to_excel([extracted_data], output_file_name)
                    if success:
                        print(f"\nPrzetwarzanie zakończone. Wyniki zapisane w {output_file_name}")
                        
                        # Pokaż ile wierszy zostanie utworzonych
                        if extracted_data.get('material_descriptions_list'):
                            print(f"Utworzono {len(extracted_data['material_descriptions_list'])} wierszy dla pozycji materiałów")
                    else:
                        print("Błąd podczas przetwarzania")
                        
                    # Pokaż wyekstrahowane dane
                    print("\nWyekstrahowane dane:")
                    for key, value in extracted_data.items():
                        if key not in ['order_items', 'material_descriptions_list']:  # Wyświetl osobno dla czytelności
                            print(f"{key}: {value}")
                    
                    if extracted_data.get('material_descriptions_list'):
                        print(f"\nZnalezione materiały ({len(extracted_data['material_descriptions_list'])}):")
                        for i, desc in enumerate(extracted_data['material_descriptions_list'], 1):
                            print(f"  {i}. {desc}")
                else:
                    print("Plik jest pusty")
                    
            except UnicodeDecodeError:
                print("Błąd dekodowania pliku. Próbuję z innym kodowaniem...")
                try:
                    with open(txt_path, 'r', encoding='latin-1') as f:
                        email_content = f.read()
                    print("Pomyślnie wczytano plik z kodowaniem latin-1")
                    # Można tutaj powtórzyć logikę przetwarzania
                except Exception as e:
                    print(f"Nie udało się wczytać pliku: {str(e)}")
            
            except Exception as e:
                print(f"Błąd podczas wczytywania pliku: {str(e)}")
    elif choice == '5':
        # Eksport do JSON
        if os.path.exists(OUTPUT_FILE):
            print(f"Eksportowanie danych z {OUTPUT_FILE} do JSON...")
            try:
                df = pd.read_excel(OUTPUT_FILE)
                data_list = df.to_dict('records')
                json_file = OUTPUT_FILE.replace('.xlsx', '.json')
                extractor.export_to_json(data_list, json_file)
                print(f"Dane zostały wyeksportowane do {json_file}")
            except Exception as e:
                print(f"Błąd podczas eksportu do JSON: {str(e)}")
        else:
            print(f"Plik {OUTPUT_FILE} nie istnieje. Najpierw przetwórz dane.")
    
    elif choice == '6':
        # Ewaluacja jakości ekstrakcji
        pdf_path = input("Podaj ścieżkę do pliku PDF do ewaluacji: ")
        if os.path.exists(pdf_path):
            print(f"Ewaluacja jakości ekstrakcji dla {pdf_path}...")
            extracted_data = extractor.extract_from_pdf(pdf_path)
            
            quality = extractor.evaluate_extraction_quality(extracted_data)
            
            print("\nStatystyki jakości ekstrakcji:")
            print(f"- Kompletność: {quality['completeness']*100:.1f}%")
            print(f"- Średnia pewność: {quality['confidence']*100:.1f}%")
            print(f"- Wszystkie wymagane pola: {'Tak' if quality['required_fields_present'] else 'Nie'}")
            
            if quality['warnings']:
                print("\nOstrzeżenia:")
                for warning in quality['warnings']:
                    print(f"  - {warning}")
            
            # Pokaż wyekstrahowane dane
            print("\nWyekstrahowane dane:")
            for key, value in extracted_data.items():
                if key != 'order_items' and isinstance(value, (str, int, float)):  # Wyświetl proste wartości
                    print(f"{key}: {value}")
        else:
            print("Plik nie istnieje")
    
    else:
        print("Nieprawidłowy wybór")

if __name__ == "__main__":
    main()