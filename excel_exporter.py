import pandas as pd
import json
import logging
from datetime import datetime
from openpyxl.styles import PatternFill
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handles Excel export with formatting"""
    
    def save_to_excel(self, data_list: List[Dict[str, Any]], output_file: str) -> bool:
        """Save extracted data to Excel with formatting"""
        try:
            # Prepare data for DataFrame
            rows = self._prepare_rows(data_list)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Extracted_Data', index=False)
                
                # Add formatting
                self._apply_formatting(writer, df)
            
            logger.info(f"Data saved to Excel: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")
            return False
    
    def _prepare_rows(self, data_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare data rows for Excel export"""
        rows = []
        
        for data in data_list:
            # Handle multiple material descriptions
            material_descriptions = []
            if data.get('material_description'):
                material_descriptions = [desc.strip() for desc in data.get('material_description', '').split(';') if desc.strip()]
            
            if not material_descriptions:
                material_descriptions = ['']
            
            # Handle order_items as JSON string
            order_items_str = ""
            if 'order_items' in data:
                if isinstance(data['order_items'], list):
                    order_items_str = json.dumps(data['order_items'], ensure_ascii=False)
                else:
                    order_items_str = str(data['order_items'])
            
            # Create separate row for each material description
            for i, material_desc in enumerate(material_descriptions):
                row = self._create_row(data, material_desc, order_items_str, i)
                rows.append(row)
        
        return rows
    
    def _create_row(self, data: Dict[str, Any], material_desc: str, order_items_str: str, index: int) -> Dict[str, Any]:
        """Create a single row for Excel"""
        # Get corresponding codes and quantities
        material_code = self._get_indexed_value(data, 'material_code', 'material_codes', index)
        colour_code = self._get_indexed_value(data, 'colour_code', 'colour_codes', index)
        quantity = self._get_indexed_value(data, 'quantity', 'quantities', index)
        
        return {
            # Required fields (green)
            'Customer_Name': data.get('customer_name', ''),
            'PO_Number': data.get('po_number', ''),
            'Material_Code': material_code,
            'Material_Description': material_desc,
            'Shipping_Street': data.get('shipping_street', ''),
            'Shipping_Postcode': data.get('shipping_postcode', ''),
            'Colour_Code': colour_code,
            'Fan_Code': data.get('fan_code', ''),
            'Shipping_Condition': data.get('shipping_condition', ''),
            
            # Additional fields
            'Quantity': quantity,
            'Project_Number': data.get('project_number', ''),
            'Date': data.get('date', ''),
            'Document_Type': data.get('document_type', ''),
            'Email_Type': data.get('email_type', ''),
            'Confidence': data.get('confidence', ''),
            'Reference_Number': data.get('reference_number', ''),
            'Order_Items': order_items_str if index == 0 else '',
            'Source_File': data.get('source_file', ''),
            'Processing_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Extraction_Method': 'Transformer' if data.get('use_transformers') else 'Traditional',
            'Item_Number': index + 1,
            'Total_Items': len(data.get('material_descriptions_list', ['']))
        }
    
    def _get_indexed_value(self, data: Dict[str, Any], single_key: str, multiple_key: str, index: int) -> str:
        """Get value by index from single or multiple values"""
        # First check if we have multiple values (quantities array)
        if data.get(multiple_key) and isinstance(data[multiple_key], list):
            if index < len(data[multiple_key]):
                return str(data[multiple_key][index])
            else:
                # If index is out of range, return empty string
                return ""
        # If we have a single value, show it on ALL rows
        elif data.get(single_key):
            return str(data[single_key])
        # No value found
        else:
            return ""
    
    def _apply_formatting(self, writer, df):
        """Apply formatting to Excel worksheet"""
        workbook = writer.book
        worksheet = writer.sheets['Extracted_Data']
        
        # Define colors
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        # Format headers
        required_fields = ['Customer_Name', 'PO_Number', 'Material_Code', 
                          'Material_Description', 'Shipping_Street', 
                          'Shipping_Postcode', 'Colour_Code', 
                          'Fan_Code', 'Shipping_Condition']
        
        helper_fields = ['Item_Number', 'Total_Items', 'Quantity']
        
        for i, column in enumerate(df.columns):
            cell = worksheet.cell(row=1, column=i+1)
            
            if column in required_fields:
                cell.fill = green_fill
            elif column in helper_fields:
                cell.fill = yellow_fill
            else:
                cell.fill = light_blue_fill
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width